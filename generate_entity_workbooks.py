#!/usr/bin/env python3
"""Generate individual reserve funding Excel workbooks for each Scarborough Glen entity."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from copy import copy

# ---------------------------------------------------------------------------
# Data from becht_report.md — all allocations, components, and disbursements
# ---------------------------------------------------------------------------

ENTITIES = {
    "HOA": {
        "name": "Scarborough Glen HOA",
        "filename": "HOA_Reserve_Plan.xlsx",
        "units": 136,
        "starting_balance": 59717,
        "baseline_contribution": 64563,
        "total_replacement_cost": 945673,
        "per_unit_floor": 350,  # ~5% of $6,953 per-unit replacement cost
        "components": [
            # (Component Name, Becht Total, Share%, Entity Allocated, Replacement Year, Inflated Cost)
            ("Aerator", 5500, 100.0, 5500, 2036, 7392),
            ("Asphalt Crack Filling", 15000, 100.0, 15000, "recurring", None),
            ("Asphalt Paving (Roads)", 330000, 100.0, 330000, 2036, 443487),
            ("Bathroom Refurbishment (Clubhouse)", 10000, 100.0, 10000, 2033, 12299),
            ("Clubhouse Furniture", 15000, 100.0, 15000, 2036, 20159),
            ("Concrete Pool Apron", 78250, 100.0, 78250, 2036, 105162),
            ("Concrete Pool Coping", 9900, 100.0, 9900, 2031, 11477),
            ("Concrete Sidewalks (HOA share)", 20103, 57.11, 20103, 2033, 24723),
            ("Curbing, Concrete", 35000, 100.0, 35000, 2036, 47037),
            ("Entrance Sign - Large", 20000, 100.0, 20000, 2046, 36122),
            ("Entrance Sign - Small", 12000, 100.0, 12000, 2046, 21673),
            ("Fence, Vinyl Stockade", 42500, 100.0, 42500, 2043, 70246),
            ("Fence, Wood Split Rail", 122500, 100.0, 122500, 2041, 190852),
            ("Fire Alarm Control Panel", 8000, 100.0, 8000, 2033, 9839),
            ("Fitness Room Refurbishment", 5000, 100.0, 5000, 2033, 6149),
            ("Gutters - Clubhouse", 2500, 100.0, 2500, 2028, 2652),
            ("Hot Water Heater", 2500, 100.0, 2500, "recurring", None),
            ("Kitchen Refurbishment", 3500, 100.0, 3500, 2033, 4305),
            ("Leaders - Clubhouse", 500, 100.0, 500, 2032, 597),
            ("Lights, Entry Clubhouse", 2700, 100.0, 2700, 2036, 3629),
            ("Lights, Entrance Sign", 1500, 100.0, 1500, 2036, 2016),
            ("Lights, Recessed", 5500, 100.0, 5500, 2033, 6764),
            ("Lights, Street", 83200, 100.0, 83200, 2033, 102325),
            ("Mailboxes", 20400, 100.0, 20400, 2049, 40261),
            ("Pool Filter System", 10000, 100.0, 10000, 2036, 13439),
            ("Roof, Shingles - Clubhouse", 15620, 100.0, 15620, 2032, 18650),
            ("Seal Coating", 16500, 100.0, 16500, "recurring", None),
            ("Siding, Vinyl - Clubhouse", 30000, 100.0, 30000, 2048, 57482),
            ("Skylights - Clubhouse", 4500, 100.0, 4500, 2032, 5373),
            ("Stop Signs", 2500, 100.0, 2500, 2046, 4515),
            ("Street Signs", 2000, 100.0, 2000, 2046, 3612),
            ("Windows, Dbl Hung Double - Clubhouse", 4000, 100.0, 4000, 2053, None),
            ("Windows, Dbl Hung Single - Clubhouse", 2000, 100.0, 2000, 2053, None),
            ("Windows, Transom", 1500, 100.0, 1500, 2053, None),
        ],
        # Year -> total disbursement (inflated) from Appendix D
        "disbursements": {
            2028: 36067,
            2031: 11477,
            2032: 24620,
            2033: 208222,
            2036: 642321,
            2038: 44915,
            2041: 190852,
            2043: 126440,
            2046: 65922,
            2048: 117838,
            2049: 40261,
        },
        # Disbursement detail: year -> [(component, amount)]
        "disbursement_detail": {
            2028: [("Asphalt Crack Filling", 15915), ("Seal Coating", 17500), ("Gutters - Clubhouse", 2652)],
            2031: [("Concrete Pool Coping", 11477)],
            2032: [("Roof - Clubhouse", 18650), ("Skylights - Clubhouse", 5373), ("Leaders - Clubhouse", 597)],
            2033: [("Lights, Street", 102325), ("Bathroom Refurbishment", 12299), ("Concrete Sidewalks (HOA)", 24723),
                   ("Fire Alarm Control Panel", 9839), ("Lights, Recessed", 6764), ("Fitness Room Refurb", 6149),
                   ("Hot Water Heater", 3070), ("Kitchen Refurbishment", 4305), ("Asphalt Crack Filling", 18444),
                   ("Seal Coating", 20304)],
            2036: [("Asphalt Paving (Roads)", 443487), ("Concrete Pool Apron", 105162), ("Curbing, Concrete", 47037),
                   ("Clubhouse Furniture", 20159), ("Pool Filter System", 13439), ("Aerator", 7392),
                   ("Lights, Entry Clubhouse", 3629), ("Lights, Entrance Sign", 2016)],
            2038: [("Asphalt Crack Filling", 21380), ("Seal Coating", 23535)],
            2041: [("Fence, Wood Split Rail", 190852)],
            2043: [("Fence, Vinyl Stockade", 70246), ("Asphalt Crack Filling", 24769), ("Seal Coating", 27274),
                   ("Hot Water Heater", 4151)],
            2046: [("Entrance Sign - Large", 36122), ("Entrance Sign - Small", 21673),
                   ("Stop Signs", 4515), ("Street Signs", 3612)],
            2048: [("Siding, Vinyl - Clubhouse", 57482), ("Asphalt Crack Filling", 28702), ("Seal Coating", 31654)],
            2049: [("Mailboxes", 40261)],
        },
        "omitted_items": [
            ("Playground Equipment", 27000, "Will need replacement; Falcon estimated $27,000"),
            ("Pool Shell Resurfacing", 22620, "Recurring ~10-year cycle; Falcon estimated $22,620/cycle"),
            ("Pool Fence, Aluminum", 18425, "Not in Becht; Falcon estimated $18,425"),
            ("Board-on-Board Fence", 37350, "Property-line fencing; Falcon estimated $37,350"),
            ("Roadway Granite Block Entry", 37875, "Decorative entry feature; Falcon estimated $37,875"),
            ("Clubhouse HVAC Split System", 9000, "Critical building system; Falcon estimated $9,000"),
            ("Clubhouse Restroom/Lockers", 30000, "Large item; Falcon estimated $30,000"),
            ("Pool Cover", 4000, "Recurring; Falcon estimated $4,000"),
            ("Pool Pump", 2500, "Recurring; Falcon estimated $2,500"),
            ("Pool Furniture Fund", 12000, "Recurring; Falcon estimated $12,000"),
            ("Pool Chlorination Equipment", 1500, "Recurring; Falcon estimated $1,500"),
            ("Clubhouse Flooring (carpet + tile)", 8280, "Falcon estimated $8,280 combined"),
            ("Clubhouse Door, Main Entry", 3500, "Falcon estimated $3,500"),
            ("Guard Rail", 10880, "Falcon estimated $10,880"),
            ("Vinyl on Masonry Wall Fence", 13950, "Falcon estimated $13,950"),
            ("Fire Suppression System", 7500, "Falcon estimated $7,500"),
            ("Irrigation Repair Fund", 5000, "Recurring; Falcon estimated $5,000"),
            ("CCTV Security", 2000, "Falcon estimated $2,000"),
            ("Key Fob Entry System", 1500, "Falcon estimated $1,500"),
        ],
    },
    "Condo_I": {
        "name": "Scarborough Glen Condo Association I",
        "filename": "Condo_I_Reserve_Plan.xlsx",
        "units": 18,
        "starting_balance": 29418,
        "baseline_contribution": 14383,
        "total_replacement_cost": 455954,
        "per_unit_floor": 1250,  # ~5% of $25,331 per-unit replacement cost
        "components": [
            ("Siding, Vinyl (Condo I share)", 1800000, 8.23, 148140, 2049, 292367),
            ("Deck Replacement, Wood (Condo I share)", 440000, 24.97, 109868, 2033, 135123),
            ("Deck Replacement, Composite (Condo I share)", 360000, 24.97, 89892, 2053, None),
            ("Roof, Shingles - Townhouses (Condo I share)", 1150988, 7.28, 83792, 2050, None),
            ("Chimney Chase Covers (Condo I share)", 178800, 10.00, 17880, 2031, 20728),
            ("Concrete Sidewalks (Condo I share)", 35200, 18.13, 6382, 2033, 7849),
        ],
        "disbursements": {
            2031: 20728,
            2033: 142972,
            2049: 292367,
        },
        "disbursement_detail": {
            2031: [("Chimney Chase Covers", 20728)],
            2033: [("Deck Replacement, Wood", 135123), ("Concrete Sidewalks", 7849)],
            2049: [("Siding, Vinyl", 292367)],
        },
        "omitted_items": [
            ("Entry Stoops (8 units)", 12000, "Concrete steps/landings at unit entries; Falcon estimated $12,000"),
            ("Wood Trim Replacement (recurring 3-yr cycle)", 5000, "Recurring maintenance; Falcon estimated $5,000/cycle"),
            ("Mailbox Hut Restoration", 1500, "Falcon estimated $1,500"),
            ("Walkway Pavers at Mailbox", 2200, "Falcon estimated $2,200"),
        ],
    },
    "Condo_II": {
        "name": "Scarborough Glen Condo Association II",
        "filename": "Condo_II_Reserve_Plan.xlsx",
        "units": 11,
        "starting_balance": 27319,
        "baseline_contribution": 18686,
        "total_replacement_cost": 423432,
        "per_unit_floor": 2000,  # ~5% of $38,494 per-unit replacement cost
        "components": [
            ("Siding, Wood (Condo II share)", 222000, 49.31, 109468, 2038, 156074),
            ("Deck Replacement, Wood (Condo II share)", 440000, 18.48, 81312, 2033, 100003),
            ("Deck Replacement, Composite (Condo II share)", 360000, 18.48, 66528, 2053, None),
            ("Roof, Shingles - Townhouses (Condo II share)", 1150988, 9.28, 106812, 2050, None),
            ("Asphalt Driveways (Condo II share)", 308000, 12.22, 37638, 2036, 50582),
            ("Chimney Chase Covers (Condo II share)", 178800, 11.00, 19668, 2031, 22801),
            ("Concrete Sidewalks (Condo II share)", 35200, 5.70, 2006, 2033, 2468),
        ],
        "disbursements": {
            2031: 22801,
            2033: 102471,
            2036: 50582,
            2038: 156074,
        },
        "disbursement_detail": {
            2031: [("Chimney Chase Covers", 22801)],
            2033: [("Deck Replacement, Wood", 100003), ("Concrete Sidewalks", 2468)],
            2036: [("Asphalt Driveways", 50582)],
            2038: [("Siding, Wood", 156074)],
        },
        "omitted_items": [
            ("Entry Stoops (11 units)", 44000, "Concrete steps/landings at unit entries; Falcon estimated $44,000"),
            ("Driveway Seal Coat (recurring 5-yr cycle)", 1400, "Recurring maintenance; Falcon estimated $1,400/cycle"),
        ],
    },
    "Condo_III": {
        "name": "Scarborough Glen Condo Association III",
        "filename": "Condo_III_Reserve_Plan.xlsx",
        "units": 9,
        "starting_balance": 24800,
        "baseline_contribution": 17450,
        "total_replacement_cost": 384386,
        "per_unit_floor": 2000,  # ~5% of $42,710 per-unit replacement cost
        "components": [
            ("Siding, Wood (Condo III share)", 222000, 50.69, 112532, 2038, 160442),
            ("Deck Replacement, Wood (Condo III share)", 440000, 19.75, 86900, 2033, 106876),
            ("Deck Replacement, Composite (Condo III share)", 360000, 19.75, 71100, 2053, None),
            ("Roof, Shingles - Townhouses (Condo III share)", 1150988, 6.98, 80339, 2050, None),
            ("Asphalt Driveways (Condo III share)", 308000, 5.13, 15800, 2036, 21234),
            ("Chimney Chase Covers (Condo III share)", 178800, 9.00, 16092, 2031, 18655),
            ("Concrete Sidewalks (Condo III share)", 35200, 4.61, 1623, 2033, 1996),
        ],
        "disbursements": {
            2031: 18655,
            2033: 108871,
            2036: 21234,
            2038: 160442,
        },
        "disbursement_detail": {
            2031: [("Chimney Chase Covers", 18655)],
            2033: [("Deck Replacement, Wood", 106876), ("Concrete Sidewalks", 1996)],
            2036: [("Asphalt Driveways", 21234)],
            2038: [("Siding, Wood", 160442)],
        },
        "omitted_items": [
            ("Entry Stoops (9 units)", 31500, "Concrete steps/landings at unit entries; Falcon estimated $31,500"),
            ("Wood Trim Replacement (recurring 5-yr cycle)", 7500, "Recurring maintenance; Falcon estimated $7,500/cycle"),
            ("Driveway Seal Coat (recurring 5-yr cycle)", 588, "Recurring maintenance; Falcon estimated $588/cycle"),
        ],
    },
    "Condo_IV": {
        "name": "Scarborough Glen Condo Association IV",
        "filename": "Condo_IV_Reserve_Plan.xlsx",
        "units": 98,
        "starting_balance": 207172,
        "baseline_contribution": 99619,
        "total_replacement_cost": 3211033,
        "per_unit_floor": 1500,  # ~5% of $32,765 per-unit replacement cost
        "components": [
            ("Siding, Vinyl (Condo IV share)", 1800000, 91.77, 1651860, 2049, 3260083),
            ("Roof, Shingles - Townhouses (Condo IV share)", 1150988, 76.46, 880045, 2050, None),
            ("Asphalt Driveways (Condo IV share)", 308000, 82.66, 254562, 2036, 342150),
            ("Deck Replacement, Wood (Condo IV share)", 440000, 36.79, 161876, 2033, 199087),
            ("Deck Replacement, Composite (Condo IV share)", 360000, 36.79, 132444, 2053, None),
            ("Chimney Chase Covers (Condo IV share)", 178800, 70.00, 125160, 2031, 145095),
            ("Concrete Sidewalks (Condo IV share)", 35200, 14.45, 5086, 2033, 6256),
        ],
        "disbursements": {
            2031: 145095,
            2033: 205342,
            2036: 342150,
            2049: 3260083,
        },
        "disbursement_detail": {
            2031: [("Chimney Chase Covers", 145095)],
            2033: [("Deck Replacement, Wood", 199087), ("Concrete Sidewalks", 6256)],
            2036: [("Asphalt Driveways", 342150)],
            2049: [("Siding, Vinyl", 3260083)],
        },
        "omitted_items": [
            ("Entry Stoops/Porches (recurring)", 10000, "End unit porches on 5-yr cycle; Falcon estimated $10,000/cycle"),
            ("Privacy Fence, Vinyl", 7735, "Condo-specific fencing; Falcon estimated $7,735"),
            ("Retaining Wall, Wood Tie Wells", 15000, "Structural/safety; Falcon estimated $15,000"),
            ("Driveway Seal Coat (recurring 5-yr cycle)", 9473, "Recurring maintenance; Falcon estimated $9,473/cycle"),
        ],
    },
}

YEARS = list(range(2026, 2051))

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LABEL_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="2F5496")
NORMAL_FONT = Font(name="Calibri", size=11)
SMALL_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FONT = Font(name="Calibri", color="9C0006", bold=True)
CURRENCY_FMT = '#,##0'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, fmt=None):
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt


def auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_instructions_sheet(ws, entity):
    ws.sheet_properties.tabColor = "2F5496"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    row = 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2, value=entity["name"]).font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=2, value="Capital Reserve Funding Plan").font = SECTION_FONT
    row += 2

    ws.cell(row=row, column=2, value="Data Sources:").font = LABEL_FONT
    row += 1
    ws.cell(row=row, column=2,
            value="Based on Becht Engineering Capital Reserve Study (Project 25-1140, February 2026)").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=2,
            value="Entity allocations derived from Falcon Group Report (2022) per-entity quantity ratios").font = NORMAL_FONT
    row += 2

    # Overview
    ws.cell(row=row, column=2, value="Overview").font = SECTION_FONT
    row += 1
    info = [
        f"Number of units: {entity['units']}",
        f"Total replacement cost (current dollars): ${entity['total_replacement_cost']:,.0f}",
        f"Baseline Year-1 annual contribution: ${entity['baseline_contribution']:,.0f}",
        f"Starting reserve balance (2026): ${entity['starting_balance']:,.0f}",
        "Planning horizon: 2026-2050 (25 years)",
    ]
    for line in info:
        ws.cell(row=row, column=2, value=line).font = NORMAL_FONT
        row += 1
    row += 1

    # Funding model
    ws.cell(row=row, column=2, value="Funding Model Assumptions").font = SECTION_FONT
    row += 1
    assumptions = [
        "Annual contributions grow at a configurable rate (default 3%) to keep pace with inflation.",
        "Reserve funds earn interest at a configurable rate (default 1%) annually.",
        "Disbursements (expenditures) occur in the year projected by the Becht engineering study.",
        "All disbursement amounts include 3% annual inflation applied to current replacement costs.",
    ]
    for line in assumptions:
        ws.cell(row=row, column=2, value=line).font = NORMAL_FONT
        row += 1
    row += 1

    # Schedule sheet guide
    ws.cell(row=row, column=2, value="Reserve Schedule Sheet — Column Guide").font = SECTION_FONT
    row += 1
    columns_desc = [
        ("Year", "Calendar year (2026-2050)"),
        ("Opening Balance", "Reserve fund balance at start of year (= prior year's Closing Balance)"),
        ("Annual Contribution", "Amount contributed this year; grows by the Growth Rate each year"),
        ("Disbursement", "Total expenditures in this year for component replacements"),
        ("Interest Earned", "Interest on the Opening Balance at the configured Interest Rate"),
        ("Closing Balance", "Opening + Contribution - Disbursement + Interest"),
        ("Monthly Cost/Unit", "Annual Contribution / Number of Units / 12 months"),
        ("Floor Check", "Shows 'BELOW FLOOR' if Closing Balance < Total Floor Amount (Per Unit x Units); blank if OK"),
    ]
    for col_name, desc in columns_desc:
        ws.cell(row=row, column=2, value=col_name).font = LABEL_FONT
        ws.cell(row=row, column=3, value=desc).font = NORMAL_FONT
        row += 1
    row += 1

    # Floor amount explanation
    ws.cell(row=row, column=2, value="Floor Amount Per Unit (Minimum Balance Threshold)").font = SECTION_FONT
    row += 1
    per_unit_floor = entity["per_unit_floor"]
    total_floor = per_unit_floor * entity["units"]
    floor_lines = [
        "The Floor Amount is set PER UNIT and multiplied by your unit count to get the total minimum balance.",
        f"Default: ${per_unit_floor:,}/unit x {entity['units']} units = ${total_floor:,} total floor.",
        f"This default is approximately 5% of your per-unit replacement cost "
        f"(${entity['total_replacement_cost'] / entity['units']:,.0f}/unit).",
        "Set to $0/unit for baseline (no cushion — fund may hit zero).",
        f"A 10% threshold would be ~${int(entity['total_replacement_cost'] * 0.10 / entity['units'] / 50) * 50:,}/unit "
        f"(${entity['total_replacement_cost'] * 0.10:,.0f} total).",
        "Higher floor = more protection against surprises but higher annual contributions needed.",
        "The yellow 'Floor Amount Per Unit' cell on the Reserve Schedule sheet is where you set this.",
        "Total Floor Amount is calculated automatically (Per Unit x Units).",
    ]
    for line in floor_lines:
        ws.cell(row=row, column=2, value=line).font = NORMAL_FONT
        row += 1
    row += 1

    # Components summary
    ws.cell(row=row, column=2, value="Components Covered by This Plan").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=2, value="See the 'Components' sheet for full detail. Summary:").font = NORMAL_FONT
    row += 2
    # Header row
    ws.cell(row=row, column=2, value="Component").font = LABEL_FONT
    ws.cell(row=row, column=3, value="Allocated Cost").font = LABEL_FONT
    ws.cell(row=row, column=4, value="Replacement Year").font = LABEL_FONT
    row += 1
    for comp in entity["components"]:
        name, _, share_pct, allocated, repl_year, _ = comp
        ws.cell(row=row, column=2, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=allocated).font = NORMAL_FONT
        ws.cell(row=row, column=3).number_format = CURRENCY_FMT
        year_val = repl_year if isinstance(repl_year, int) else None
        if year_val:
            ws.cell(row=row, column=4, value=year_val).font = NORMAL_FONT
        row += 1
    row += 1

    # Omitted items
    if entity.get("omitted_items"):
        ws.cell(row=row, column=2,
                value="Items NOT Included (Budget Separately)").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=2,
                value="The following items from the Falcon 2022 report are NOT in the Becht 2026 study "
                      "and are therefore NOT included in this funding schedule. Consider budgeting "
                      "for these items separately.").font = NORMAL_FONT
        row += 2
        # Header
        ws.cell(row=row, column=2, value="Item").font = LABEL_FONT
        ws.cell(row=row, column=3, value="Falcon Est. Cost").font = LABEL_FONT
        ws.cell(row=row, column=4, value="Notes").font = LABEL_FONT
        row += 1
        total_omitted = 0
        for item_name, cost, notes in entity["omitted_items"]:
            ws.cell(row=row, column=2, value=item_name).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=cost).font = NORMAL_FONT
            ws.cell(row=row, column=3).number_format = CURRENCY_FMT
            ws.cell(row=row, column=4, value=notes).font = SMALL_FONT
            total_omitted += cost
            row += 1
        ws.cell(row=row, column=2, value="TOTAL OMITTED").font = LABEL_FONT
        ws.cell(row=row, column=3, value=total_omitted).font = LABEL_FONT
        ws.cell(row=row, column=3).number_format = CURRENCY_FMT
        row += 2

    ws.cell(row=row, column=2,
            value="Prepared June 2026. Review and update annually.").font = SMALL_FONT


def build_schedule_sheet(ws, entity):
    ws.sheet_properties.tabColor = "548235"
    units = entity["units"]

    # --- Input area ---
    row = 1
    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value="RESERVE FUNDING INPUTS").font = SECTION_FONT

    per_unit_floor = entity["per_unit_floor"]

    labels_values = [
        ("Starting Balance (2026):", entity["starting_balance"], CURRENCY_FMT, True),
        ("Year-1 Annual Contribution:", entity["baseline_contribution"], CURRENCY_FMT, True),
        ("Annual Contribution Growth Rate:", 0.03, PCT_FMT, True),
        ("Interest Rate on Reserves:", 0.01, PCT_FMT, True),
        ("Floor Amount Per Unit:", per_unit_floor, CURRENCY_FMT, True),
        ("Number of Units:", units, "#,##0", True),
        (None, None, None, False),  # row 8: computed total floor
    ]

    for i, (label, val, fmt, is_input) in enumerate(labels_values):
        r = 2 + i
        if label is None:
            continue
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.font = LABEL_FONT
        c.number_format = fmt
        if is_input:
            c.fill = INPUT_FILL
            c.border = THIN_BORDER

    # Row 8: Computed Total Floor = Per Unit Floor * Units (formula, not editable)
    ws.cell(row=8, column=1, value="Total Floor Amount:").font = LABEL_FONT
    c = ws.cell(row=8, column=2)
    c.value = "=$B$6*$B$7"
    c.font = LABEL_FONT
    c.number_format = CURRENCY_FMT
    c.border = THIN_BORDER

    # Named references for formulas (using $B$ references)
    # Row 2: Starting Balance -> B2
    # Row 3: Year-1 Contribution -> B3
    # Row 4: Growth Rate -> B4
    # Row 5: Interest Rate -> B5
    # Row 6: Floor Per Unit -> B6
    # Row 7: Units -> B7
    # Row 8: Total Floor (formula) -> B8

    ws.cell(row=9, column=1,
            value="Edit the yellow cells above to customize your plan. Total Floor = Per Unit Floor x Units.").font = SMALL_FONT

    # --- Schedule table header ---
    header_row = 11
    headers = [
        "Year", "Opening Balance", "Annual Contribution", "Disbursement",
        "Interest Earned", "Closing Balance", "Monthly Cost/Unit", "Floor Check"
    ]

    for col_idx, hdr in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=hdr)
    style_header_row(ws, header_row, len(headers))

    # --- Data rows ---
    first_data_row = header_row + 1
    disbursements = entity["disbursements"]

    for i, year in enumerate(YEARS):
        r = first_data_row + i
        disb = disbursements.get(year, 0)

        # Col A: Year
        cell_year = ws.cell(row=r, column=1, value=year)
        style_data_cell(cell_year)

        # Col B: Opening Balance
        if i == 0:
            ws.cell(row=r, column=2).value = f"=$B$2"  # starting balance
        else:
            ws.cell(row=r, column=2).value = f"=F{r - 1}"  # prior closing
        cell_ob = ws.cell(row=r, column=2)
        cell_ob.number_format = CURRENCY_FMT
        cell_ob.border = THIN_BORDER

        # Col C: Annual Contribution
        if i == 0:
            ws.cell(row=r, column=3).value = f"=$B$3"
        else:
            ws.cell(row=r, column=3).value = f"=C{r - 1}*(1+$B$4)"
        cell_ac = ws.cell(row=r, column=3)
        cell_ac.number_format = CURRENCY_FMT
        cell_ac.border = THIN_BORDER

        # Col D: Disbursement (hard-coded)
        cell_d = ws.cell(row=r, column=4, value=disb)
        style_data_cell(cell_d, CURRENCY_FMT)

        # Col E: Interest Earned
        ws.cell(row=r, column=5).value = f"=B{r}*$B$5"
        cell_ie = ws.cell(row=r, column=5)
        cell_ie.number_format = CURRENCY_FMT
        cell_ie.border = THIN_BORDER

        # Col F: Closing Balance
        ws.cell(row=r, column=6).value = f"=B{r}+C{r}-D{r}+E{r}"
        cell_cb = ws.cell(row=r, column=6)
        cell_cb.number_format = CURRENCY_FMT
        cell_cb.border = THIN_BORDER

        # Col G: Monthly Cost/Unit
        ws.cell(row=r, column=7).value = f"=C{r}/$B$7/12"
        cell_mu = ws.cell(row=r, column=7)
        cell_mu.number_format = '$#,##0'
        cell_mu.border = THIN_BORDER

        # Col H: Floor Check
        ws.cell(row=r, column=8).value = f'=IF(F{r}<$B$8,"BELOW FLOOR","")'
        cell_fc = ws.cell(row=r, column=8)
        cell_fc.border = THIN_BORDER
        cell_fc.alignment = Alignment(horizontal="center")

    # --- Totals row ---
    totals_row = first_data_row + len(YEARS)
    ws.cell(row=totals_row, column=1, value="TOTAL").font = LABEL_FONT
    ws.cell(row=totals_row, column=1).border = THIN_BORDER

    last_data = first_data_row + len(YEARS) - 1
    for col in [3, 4, 5]:
        col_letter = get_column_letter(col)
        ws.cell(row=totals_row, column=col).value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data})"
        ws.cell(row=totals_row, column=col).font = LABEL_FONT
        ws.cell(row=totals_row, column=col).number_format = CURRENCY_FMT
        ws.cell(row=totals_row, column=col).border = THIN_BORDER

    # Apply conditional formatting for below-floor years
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"F{first_data_row}:F{last_data}",
        CellIsRule(operator="lessThan", formula=["$B$8"], fill=WARN_FILL, font=WARN_FONT),
    )

    # Column widths
    col_widths = [8, 18, 20, 18, 16, 18, 18, 16]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Freeze panes
    ws.freeze_panes = f"A{first_data_row}"

    # --- Disbursement detail section ---
    detail_start_col = 10  # Column J
    ws.cell(row=header_row - 1, column=detail_start_col, value="DISBURSEMENT DETAIL").font = SECTION_FONT

    # Collect all component names that appear in disbursements
    all_detail_components = set()
    for year_details in entity.get("disbursement_detail", {}).values():
        for comp_name, _ in year_details:
            all_detail_components.add(comp_name)
    all_detail_components = sorted(all_detail_components)

    if all_detail_components:
        # Headers
        for j, comp_name in enumerate(all_detail_components):
            col = detail_start_col + j
            cell = ws.cell(row=header_row, column=col, value=comp_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Data
        detail_data = entity.get("disbursement_detail", {})
        for i, year in enumerate(YEARS):
            r = first_data_row + i
            year_items = {name: amt for name, amt in detail_data.get(year, [])}
            for j, comp_name in enumerate(all_detail_components):
                col = detail_start_col + j
                val = year_items.get(comp_name)
                if val:
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.number_format = CURRENCY_FMT
                    cell.border = THIN_BORDER

        # Detail totals
        for j in range(len(all_detail_components)):
            col = detail_start_col + j
            col_letter = get_column_letter(col)
            ws.cell(row=totals_row, column=col).value = (
                f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data})"
            )
            ws.cell(row=totals_row, column=col).font = LABEL_FONT
            ws.cell(row=totals_row, column=col).number_format = CURRENCY_FMT
            ws.cell(row=totals_row, column=col).border = THIN_BORDER


def build_components_sheet(ws, entity):
    ws.sheet_properties.tabColor = "BF8F00"

    ws.cell(row=1, column=1, value=f"{entity['name']} — Component Detail").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="All costs from Becht Engineering Capital Reserve Study (2026)").font = SMALL_FONT

    header_row = 4
    headers = [
        "Component", "Becht Community Total", "Entity Share %",
        "Entity Allocated Cost", "Replacement Year", "Inflated Cost at Replacement"
    ]
    for col, hdr in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=hdr)
    style_header_row(ws, header_row, len(headers))

    for i, comp in enumerate(entity["components"]):
        r = header_row + 1 + i
        name, becht_total, share_pct, allocated, repl_year, inflated = comp

        ws.cell(row=r, column=1, value=name)
        style_data_cell(ws.cell(row=r, column=1))

        ws.cell(row=r, column=2, value=becht_total)
        style_data_cell(ws.cell(row=r, column=2), CURRENCY_FMT)

        ws.cell(row=r, column=3, value=share_pct / 100)
        style_data_cell(ws.cell(row=r, column=3), PCT_FMT)

        ws.cell(row=r, column=4, value=allocated)
        style_data_cell(ws.cell(row=r, column=4), CURRENCY_FMT)

        if isinstance(repl_year, int):
            year_display = repl_year if repl_year <= 2050 else f"{repl_year} (beyond plan)"
        else:
            year_display = str(repl_year)
        ws.cell(row=r, column=5, value=year_display)
        style_data_cell(ws.cell(row=r, column=5))

        if inflated is not None:
            ws.cell(row=r, column=6, value=inflated)
            style_data_cell(ws.cell(row=r, column=6), CURRENCY_FMT)
        else:
            ws.cell(row=r, column=6, value="Beyond plan horizon")
            style_data_cell(ws.cell(row=r, column=6))

    # Totals
    total_row = header_row + 1 + len(entity["components"])
    ws.cell(row=total_row, column=1, value="TOTAL").font = LABEL_FONT
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    ws.cell(row=total_row, column=4, value=entity["total_replacement_cost"])
    ws.cell(row=total_row, column=4).font = LABEL_FONT
    ws.cell(row=total_row, column=4).number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=4).border = THIN_BORDER

    col_widths = [45, 22, 14, 22, 18, 26]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.freeze_panes = f"A{header_row + 1}"


COMPONENT_COLORS = [
    "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
    "70AD47", "264478", "9B57A0", "636363", "EB7E30",
    "44546A", "BF4B28", "00B0F0", "92D050", "7030A0",
    "C55A11", "2E75B6", "AFABAB", "43682B", "D63384",
    "F4B183", "8FAADC", "A9D18E", "FFD966", "B4C7E7",
    "F8CBAD", "C9C9C9", "E2F0D9", "D6DCE4", "FBE5D6",
]


def build_chart_sheet(ws, entity):
    ws.sheet_properties.tabColor = "C00000"

    detail_data = entity.get("disbursement_detail", {})
    all_components = set()
    for year_details in detail_data.values():
        for comp_name, _ in year_details:
            all_components.add(comp_name)
    all_components = sorted(all_components)

    if not all_components:
        ws.cell(row=1, column=1, value="No disbursement detail available.").font = NORMAL_FONT
        return

    # --- Color key table (top-left) ---
    ws.cell(row=1, column=1, value="COLOR KEY").font = SECTION_FONT
    ws.cell(row=2, column=1, value="Color").font = LABEL_FONT
    ws.cell(row=2, column=2, value="Component").font = LABEL_FONT
    ws.cell(row=2, column=3, value="25-Year Total").font = LABEL_FONT
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 16

    # Sum each component's total across all years for the key table
    comp_totals = {}
    for year_items in detail_data.values():
        for comp_name, amt in year_items:
            comp_totals[comp_name] = comp_totals.get(comp_name, 0) + amt

    for j, comp in enumerate(all_components):
        r = 3 + j
        color = COMPONENT_COLORS[j % len(COMPONENT_COLORS)]
        cell_color = ws.cell(row=r, column=1)
        cell_color.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_color.border = THIN_BORDER
        ws.cell(row=r, column=2, value=comp).font = NORMAL_FONT
        ws.cell(row=r, column=2).border = THIN_BORDER
        total_val = comp_totals.get(comp, 0)
        ws.cell(row=r, column=3, value=total_val).font = NORMAL_FONT
        ws.cell(row=r, column=3).number_format = CURRENCY_FMT
        ws.cell(row=r, column=3).border = THIN_BORDER

    # --- Data table for chart (to the right of the color key) ---
    data_start_col = 5  # Column E
    years_with_data = sorted(y for y in YEARS if y in detail_data)
    num_data_rows = len(years_with_data)
    num_components = len(all_components)

    ws.cell(row=1, column=data_start_col, value="CHART DATA").font = SECTION_FONT
    ws.cell(row=2, column=data_start_col, value="Year").font = LABEL_FONT
    for j, comp in enumerate(all_components):
        ws.cell(row=2, column=data_start_col + 1 + j, value=comp).font = LABEL_FONT
        ws.column_dimensions[get_column_letter(data_start_col + 1 + j)].width = 16

    for i, year in enumerate(years_with_data):
        r = 3 + i
        ws.cell(row=r, column=data_start_col, value=year)
        year_items = {name: amt for name, amt in detail_data.get(year, [])}
        for j, comp in enumerate(all_components):
            val = year_items.get(comp)
            if val:
                ws.cell(row=r, column=data_start_col + 1 + j, value=val)
                ws.cell(row=r, column=data_start_col + 1 + j).number_format = CURRENCY_FMT

    # --- Stacked bar chart ---
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = entity["name"] + " — Projected Expenditures by Year"
    chart.style = 10
    chart.width = 36
    chart.height = 20
    chart.legend = None

    # Y-axis: dollar labels with $K formatting, gridlines, title
    chart.y_axis.title = "Cost ($)"
    chart.y_axis.numFmt = '$#,##0,K'
    chart.y_axis.majorGridlines = None  # let default gridlines show
    chart.y_axis.tickLblPos = "low"
    chart.y_axis.delete = False

    # X-axis: year labels rotated vertically
    chart.x_axis.title = "Year"
    chart.x_axis.numFmt = '0'
    chart.x_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.delete = False
    # Rotate x-axis labels 270 degrees (vertical, reading bottom-to-top)
    chart.x_axis.txPr = RichText(
        p=[Paragraph(
            pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=1000)
            ),
            endParaRPr=CharacterProperties(sz=1000),
        )]
    )
    chart.x_axis.txPr.properties.rot = -5400000  # -90 degrees in 60000ths

    cats = Reference(ws, min_col=data_start_col, min_row=3, max_row=2 + num_data_rows)

    for j in range(num_components):
        col = data_start_col + 1 + j
        data = Reference(ws, min_col=col, min_row=2, max_row=2 + num_data_rows)
        chart.add_data(data, titles_from_data=True)
        series = chart.series[j]
        color = COMPONENT_COLORS[j % len(COMPONENT_COLORS)]
        series.graphicalProperties.solidFill = color

    chart.set_categories(cats)

    # Place chart below both the color key and data table
    chart_start_row = max(3 + num_components, 3 + num_data_rows) + 2
    ws.add_chart(chart, f"A{chart_start_row}")

    ws.column_dimensions[get_column_letter(data_start_col)].width = 8


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate_workbook(key, entity):
    wb = Workbook()

    # Sheet 1: Instructions
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    build_instructions_sheet(ws_instr, entity)

    # Sheet 2: Reserve Schedule
    ws_sched = wb.create_sheet("Reserve Schedule")
    build_schedule_sheet(ws_sched, entity)

    # Sheet 3: Components
    ws_comp = wb.create_sheet("Components")
    build_components_sheet(ws_comp, entity)

    # Sheet 4: Expenditure Chart
    ws_chart = wb.create_sheet("Expenditure Chart")
    build_chart_sheet(ws_chart, entity)

    filename = entity["filename"]
    wb.save(filename)
    print(f"  Created: {filename}")


def main():
    print("Generating entity reserve funding workbooks...")
    for key, entity in ENTITIES.items():
        generate_workbook(key, entity)
    print(f"\nDone. {len(ENTITIES)} workbooks created.")


if __name__ == "__main__":
    main()
